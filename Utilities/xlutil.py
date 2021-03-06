import openpyxl
from Utilities.constant import Constants

class ExcelUtil:
    
    constants = Constants()

    def setExcelFile(self, path):
        try:
            self.workbook = openpyxl.load_workbook(path)
        except:
            print('Failed to open the file')

    def getRowCount(self, sheetName):
        try:
            sheet = self.workbook[sheetName]
            return sheet.max_row
        except:
            print('Failed to get total row count')

    def getColCount(self, sheetName):
        try:
            
            sheet = self.workbook[sheetName]
            return sheet.max_column
        except:
            print('Failed to get total column count')

    def getCellData(self, sheetName, row_count, col_count):
        try:
            sheet = self.workbook[sheetName]
            return sheet.cell(row_count, col_count).value
        except:
            print("Failed to get the cell data")

    def getRowContains(self, testname, colNum, sheetname):
        rowNum = 0
        try:
            rowCount = 0
            rowCount = self.getRowCount(sheetname)

            for rowNum in range(0, rowCount):
                if self.getCellData(sheetname, rowNum, colNum) == testname:
                    break
        except:
            print('Row contains failed in checking')

    def getTestStepsCount(self, sheetname, testname, stepstart):
        try:
            rowCount = 0

            rowCount = self.getRowCount(sheetname)
            i = 0
            for i in range(stepstart, rowCount):
                if str(testname) != str(self.getCellData(
                    sheetname, 
                    i, 
                    self.constants.Col_TestCaseID)
                ):
                    return i
        except:
            print("Failed to get steps count")
            return 0

    def getObjectValue(self, sheetname, objectname):
        try:
            rowcount = self.getRowCount(sheetname)
            for i in range(0, rowcount):
                if objectname == "":
                    break
                elif str(objectname) == str(self.getCellData(sheetname, i, 0)):
                    object_value = str(self.getCellData(
                        sheetname,
                        i,
                        self.constants.Col_Object_Value)
                    )
                    return object_value
        except:
            print("Failed to get object value")

    # Return test steps count
    def getITCount(self, sheetn, testn, stepst):
        try:
            rowCount = 0
            rowCount = self.getRowCount(sheetn)
            i = 0
            for i in range(stepst, rowCount):
                if str(testn) != str(self.getCellData(
                    sheetn, 
                    i, 
                    self.constants.Col_TestCaseID)
                ):
                    return i
        except:
            self.log.error("Failed to get steps count")
            return 0
    
    # Get test case iterations count
    def getTestIterations(self, sheetname, testname):
        try:
            isSheet = self.workbook.sheetnames
            for i in range(len(isSheet)):
                if str(testname) == str(isSheet[i]):
                    iterate = 0
                    iterate = self.getITCount(sheetname, testname, 1)
                else:
                    continue
            if iterate > 0:
                return iterate
            else:
                return 2
        except Exception as e:
            return 2

    def getTestdatavalue(self, testname, data, rownum):
        try:
            self.test_work_sheet = self.workbook.sheet_by_name(testname)
            test_rows = self.test_work_sheet.nrows
            test_cols = self.test_work_sheet.ncols
            for irows in range(0, 1):
                for icols in range(0, test_cols):
                    test_data_header = self.getCellData(irows, icols, testname)
                    if test_data_header == data:
                        testdata_value = self.getCellData(
                            rownum, 
                            icols, 
                            testname
                        )
                        return testdata_value
                    else:
                        pass
        except:
            pass

    def writeData(sheetName, row_count, col_count, data):
        sheet = self.workbook.get_sheet_by_name(sheetName)
        sheet.cell(row=row_count, column=col_count).value = data
        workbook.save(file)